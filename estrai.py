from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdftypes import resolve1
from pdfminer.psparser import PSLiteral, PSKeyword
from pdfminer.utils import decode_text
from openpyxl import Workbook
import tkinter.filedialog as affo
import tkinter.filedialog as affs
import tkinter as tk
import os

window = tk.Tk()
window.geometry("600x600")
window.title("Esporta moduli Pdf in excel")
window.resizable(False, False)
window.configure(background="white")

data = {}


def decode_value(value):

    # decode PSLiteral, PSKeyword
    if isinstance(value, (PSLiteral, PSKeyword)):
        value = value.name

    # decode bytes
    if isinstance(value, bytes):
        value = decode_text(value)

    return value

def converti():
    with open(file_path, 'rb') as fp:
        
        parser = PDFParser(fp)

        doc = PDFDocument(parser)
        res = resolve1(doc.catalog)

        if 'AcroForm' not in res:            
            labelExample5 = tk.Label(window, text='nessun modulo trovato')
            labelExample5.pack()
            raise ValueError("No AcroForm Found")
        fields = resolve1(doc.catalog['AcroForm'])['Fields']  # may need further resolving
        wb = Workbook()
        ws = wb.active
        a = 0
        for f in fields:
            field = resolve1(f)
            name, values = field.get('T'), field.get('V')

            # decode name
            name = decode_text(name)

            # resolve indirect obj
            values = resolve1(values)

            # decode value(s)
            if isinstance(values, list):
                values = [decode_value(v) for v in values]
            else:
                values = decode_value(values)

            data.update({name: values})
            ws.cell(row=a+1, column=1).value = name
            ws.cell(row=a+1, column=2).value = values
            a = a+1
            print(name, values)
        wb.save(file_exceld)
        labelExample5 = tk.Label(window, text='esportato con successo')
        labelExample5.pack()
labelExample = tk.Label(window, text="File di origine")
labelExample.pack()
file_path = affo.askopenfilename(filetypes =[('Pdf', '*.pdf')])
labelExample2 = tk.Label(window, text=file_path)
labelExample2.pack()
file_exceld = affs.asksaveasfilename(defaultextension=".xlsx", title="Salva come", filetypes=[("Excel Files", "*.xlsx")])
labelExample3 = tk.Label(window, text="File di destinazione")
labelExample3.pack()
labelExample4 = tk.Label(window, text=file_exceld)
labelExample4.pack()
second_button = tk.Button(text="Converti in excel", command=converti)
second_button.pack()
if __name__ == "__main__":
    window.mainloop()